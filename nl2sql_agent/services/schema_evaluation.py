"""Thread-safe orchestration for the read-only Schema golden-set evaluation."""

from __future__ import annotations

import threading
import time
import json
import uuid
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook

from nl2sql_agent.eval.run_schema_golden_eval import (
    DEFAULT_CASES,
    run_golden_evaluation,
    run_online_shadow_evaluation,
)


JSON_COLUMNS = {
    "tags", "data_scope", "query_intent", "expected_tables", "expected_columns",
    "forbidden_tables", "expected_plan_tables", "expected_joins",
}
REQUIRED_COLUMNS = {"id", "question", "expected_tables"}


class SchemaEvaluationService:
    def __init__(self, data_dir: str | Path | None = None) -> None:
        self._lock = threading.Lock()
        self._reports: dict[str, dict[str, Any]] = {}
        self._data_dir = Path(data_dir or DEFAULT_CASES.parents[2] / "data" / "evaluation")
        self._data_dir.mkdir(parents=True, exist_ok=True)
        self._active_path = self._data_dir / "active_dataset.json"

    def _builtin_payload(self) -> dict[str, Any]:
        return yaml.safe_load(DEFAULT_CASES.read_text(encoding="utf-8")) or {}

    def dataset_payload(self) -> dict[str, Any]:
        if self._active_path.exists():
            return json.loads(self._active_path.read_text(encoding="utf-8"))
        payload = self._builtin_payload()
        payload["dataset_id"] = "builtin"
        payload["name"] = "内置 Schema 黄金集"
        return payload

    @property
    def running(self) -> bool:
        return self._lock.locked()

    def dataset_summary(self) -> dict[str, Any]:
        payload = self.dataset_payload()
        return {
            "dataset_id": payload.get("dataset_id", "builtin"),
            "name": payload.get("name", "未命名评测集"),
            "version": payload.get("version", 1),
            "description": payload.get("description", ""),
            "coverage": payload.get("coverage") or {},
            "case_count": len(payload.get("cases") or []),
            "cases": [
                {"id": item.get("id"), "question": item.get("question"), "suite": item.get("suite")}
                for item in payload.get("cases") or []
            ],
        }

    def import_xlsx(self, content: bytes, filename: str = "evaluation.xlsx") -> dict[str, Any]:
        if len(content) > 10 * 1024 * 1024:
            raise ValueError("Excel 文件不能超过 10MB")
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook["cases"] if "cases" in workbook.sheetnames else workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows)]
        except Exception as exc:
            raise ValueError(f"无法读取 Excel 文件: {exc}") from exc
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError("Excel 缺少必填列: " + ", ".join(sorted(missing)))
        cases: list[dict[str, Any]] = []
        errors: list[str] = []
        for row_number, values in enumerate(rows, start=2):
            raw = dict(zip(headers, values))
            if not any(value is not None and str(value).strip() for value in values):
                continue
            item: dict[str, Any] = {}
            for key, value in raw.items():
                if value is None or value == "":
                    continue
                if key in JSON_COLUMNS:
                    try:
                        item[key] = json.loads(str(value))
                    except json.JSONDecodeError as exc:
                        errors.append(f"第 {row_number} 行 {key} 不是合法 JSON: {exc.msg}")
                elif key == "expected_clarification":
                    item[key] = str(value).strip().lower() in {"1", "true", "yes", "是"}
                else:
                    item[key] = str(value).strip()
            if not item.get("id") or not item.get("question"):
                errors.append(f"第 {row_number} 行 id 和 question 不能为空")
            cases.append(item)
        ids = [str(item.get("id")) for item in cases]
        duplicates = sorted({value for value in ids if ids.count(value) > 1})
        if duplicates:
            errors.append("case id 重复: " + ", ".join(duplicates))
        if not cases:
            errors.append("Excel 中没有有效用例")
        if errors:
            raise ValueError("; ".join(errors[:20]))
        payload = {
            "dataset_id": uuid.uuid4().hex[:12],
            "name": Path(filename).stem,
            "version": 1,
            "description": f"上传自 {Path(filename).name}",
            "coverage": {},
            "cases": cases,
        }
        self._active_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        self._reports.clear()
        return self.dataset_summary()

    def template_xlsx(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "cases"
        headers = [
            "id", "suite", "tags", "question", "data_scope", "query_intent",
            "expected_tables", "expected_columns", "forbidden_tables",
            "expected_plan_tables", "expected_joins", "expected_clarification",
        ]
        sheet.append(headers)
        sheet.append([
            "case_001", "basic", '["示例"]', "统计订单金额", '["risk_mart"]',
            '{"query_type":"aggregation","measures":[{"text":"订单金额","role":"measure"}]}',
            '["orders"]', '["orders.amount"]', "[]", '["orders"]', "[]", False,
        ])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(
                max(14, max(len(str(cell.value or "")) for cell in column) + 2), 60
            )
        guide = workbook.create_sheet("说明")
        guide.append(["字段", "说明"])
        guide.append(["JSON 列", "tags、data_scope、query_intent、expected_* 数组列必须填写合法 JSON"])
        guide.append(["expected_joins", '格式示例：[["orders.customer_id","customers.id"]]'])
        guide.append(["必填", "id、question、expected_tables"])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _report_key(mode: str, database_id: str | None = None) -> str:
        if mode == "online_shadow":
            return f"{mode}:{database_id or ''}"
        return mode

    def status(
        self, mode: str = "baseline", database_id: str | None = None
    ) -> dict[str, Any]:
        return {
            "running": self.running,
            "dataset": self.dataset_summary(),
            "mode": mode,
            "database_id": database_id,
            "report": self._reports.get(self._report_key(mode, database_id)),
        }

    def run(
        self, mode: str = "baseline", deps=None, database_id: str | None = None,
        case_id: str | None = None,
    ) -> dict[str, Any]:
        if mode not in {"baseline", "online_shadow"}:
            raise ValueError(f"Unsupported evaluation mode: {mode}")
        if mode == "online_shadow" and deps is None:
            raise ValueError("online_shadow requires database-bound dependencies")
        if not self._lock.acquire(blocking=False):
            raise RuntimeError("Schema 评测正在运行，请稍后查看结果")
        try:
            started_at = time.time()
            payload = self.dataset_payload()
            if case_id:
                selected = [item for item in payload.get("cases") or [] if str(item.get("id")) == case_id]
                if not selected:
                    raise ValueError(f"评测用例不存在: {case_id}")
                payload = {**payload, "cases": selected, "coverage": {"single_case": case_id}}
            report = (
                run_online_shadow_evaluation(deps, payload_override=payload)
                if mode == "online_shadow"
                else run_golden_evaluation(payload_override=payload, deps_override=deps)
            )
            report.update({
                "started_at": started_at,
                "finished_at": time.time(),
                "database_id": database_id,
                "dataset_id": payload.get("dataset_id", "builtin"),
                "case_id": case_id,
            })
            report["duration_seconds"] = round(
                report["finished_at"] - started_at, 3
            )
            self._reports[self._report_key(mode, database_id)] = report
            return report
        finally:
            self._lock.release()
