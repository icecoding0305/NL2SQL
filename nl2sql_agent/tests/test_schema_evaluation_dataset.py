from io import BytesIO

from openpyxl import load_workbook

from nl2sql_agent.services.schema_evaluation import SchemaEvaluationService


def test_xlsx_template_can_be_imported_and_persisted(tmp_path):
    service = SchemaEvaluationService(tmp_path)

    content = service.template_xlsx()
    workbook = load_workbook(BytesIO(content), read_only=True)

    assert "cases" in workbook.sheetnames
    summary = service.import_xlsx(content, "my-golden-set.xlsx")
    assert summary["name"] == "my-golden-set"
    assert summary["case_count"] == 1
    assert summary["cases"][0]["id"] == "case_001"

    reloaded = SchemaEvaluationService(tmp_path)
    assert reloaded.dataset_summary()["dataset_id"] == summary["dataset_id"]


def test_xlsx_import_rejects_missing_required_columns(tmp_path):
    service = SchemaEvaluationService(tmp_path)
    workbook = load_workbook(BytesIO(service.template_xlsx()))
    sheet = workbook["cases"]
    sheet.delete_cols(7)
    output = BytesIO()
    workbook.save(output)

    try:
        service.import_xlsx(output.getvalue(), "invalid.xlsx")
    except ValueError as exc:
        assert "expected_tables" in str(exc)
    else:
        raise AssertionError("invalid workbook should be rejected")
